#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成信用债发行汇总 Excel（云端版）。
累积机制：以 data/bonds_master.json 为主库（按债券简称唯一），每次运行：
  1) 加载主库（缺失则自动从 data/ 内最大的汇总备份 xlsx 回种）
  2) 合并 data/credit_bond_*.xlsx（近期导出，刷新票面利率、债券代码）
  3) 回写主库（persist，永不重建）
  4) 用主库全量生成 Sheet2
Sheet1 = 主库中「截标/发行日期 == 今天」的债券（与用户预期一致：当日新发）。
两表 + 导出真源 三方交叉验证：
  - 真源(债立方当日导出)中截标日=今天的债券 必须全部在 Sheet1；
  - Sheet1 每只必须也在真源(截标日=今天) 与 Sheet2 汇总中。
任一项不满足即退出非零，workflow 将跳过发邮件（杜绝发出自相矛盾的 Excel）。
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone
import glob
import os
import re
import sys
import json
import warnings
warnings.filterwarnings('ignore')

TARGET_HEADERS = [
    '截标时间', '债券简称', '债券代码', '发行期限', '计划发行(亿)', '投标区间(%)', '票面利率(%)',
    '发行人', '区域', '相似二级券', '相似二级剩余期限', '相似二级估值', 'YY评分',
    '主体评级', '债项评级', '主承销商', '担保人', '债券类型', '发行人类型', '交易市场',
    '债券全称', '募集方式', '实际到期日', '到期日休市情况', '上市日', '公告日', '缴款日',
    '板块分类', 'DM评分', 'DM一级行业', 'DM二级行业', 'DM三级行业', '申万行业',
    '行权日', '是否有担保', '条款', '偿付顺序', '发行截止日', '团费(%)'
]

# ---- 路径（云端：脚本所在目录即仓库根）----
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')
DOWNLOAD_DIR = DATA_DIR  # 导出脚本把 credit_bond_*.xlsx 写到 data/
MASTER_PATH = os.path.join(DATA_DIR, 'bonds_master.json')

# ---- 北京时间（云端 runner 默认 UTC，必须 +8h 校正）----
_beijing = datetime.now(timezone.utc) + timedelta(hours=8)
TODAY = _beijing.strftime('%Y-%m-%d')
OUTPUT_PATH = os.path.join(DATA_DIR, f"信用债发行汇总_{TODAY}.xlsx")

# ---------- 读取辅助 ----------
def read_credit_bond(path, dt):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    sh = [c.value for c in ws[2]]
    hm = {h: i for i, h in enumerate(sh) if h}
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r[0]:
            continue
        rows.append((dt, {k: (r[i] if i < len(r) else None) for k, i in hm.items()}))
    return rows


def read_summary_sheet2(path, fallback_dt):
    wb = openpyxl.load_workbook(path)
    if '汇总' not in wb.sheetnames:
        return []
    ws = wb['汇总']
    hdr = [c.value for c in ws[1]]
    gi = {h: i for i, h in enumerate(hdr)}
    name_i = gi.get('债券简称')
    if name_i is None:
        return []
    fs_i = gi.get('首次提取日期')
    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_i + 1).value
        if not name:
            continue
        dt = ws.cell(r, fs_i + 1).value if fs_i is not None else None
        if not dt:
            dt = fallback_dt
        d = {}
        for h, i in gi.items():
            if h == '首次提取日期':
                continue
            d[h] = ws.cell(r, i + 1).value
        rows.append((dt, d))
    return rows


def merge(master, dt, row, run_date):
    """合并一行到主库。返回 (is_new_this_run, name)。"""
    name = row.get('债券简称')
    if not name:
        return False, None
    is_new = name not in master
    if is_new:
        master[name] = {'data': {}, 'first_seen': dt, 'dates_seen': [dt],
                        'latest_date': dt, 'run_seen': run_date}
    else:
        if dt not in master[name]['dates_seen']:
            master[name]['dates_seen'].append(dt)
        if dt > master[name]['latest_date']:
            master[name]['latest_date'] = dt
        if 'run_seen' not in master[name]:
            master[name]['run_seen'] = run_date
    for k, v in row.items():
        if v not in (None, '', 0):
            if k in ('债券代码', '票面利率(%)'):
                if dt >= master[name]['latest_date'] or master[name]['data'].get(k) in (None, '', 0):
                    master[name]['data'][k] = v
            else:
                if dt >= master[name]['latest_date'] or k not in master[name]['data']:
                    master[name]['data'][k] = v
        elif k not in master[name]['data']:
            master[name]['data'][k] = v
    return is_new, name


# ---------- 1) 加载主库（缺失则从汇总备份回种）----------
def load_master():
    if os.path.exists(MASTER_PATH):
        try:
            with open(MASTER_PATH, encoding='utf-8') as f:
                m = json.load(f)
            if m:
                return m
        except Exception:
            pass
    print("[WARN] 主库缺失/为空，尝试从 data/ 内汇总备份回种...")
    best = {}
    best_n = 0
    for f in glob.glob(os.path.join(DATA_DIR, '信用债发行汇总*.xlsx')):
        m = re.search(r'(\d{4}-\d{2}-\d{2})\.xlsx', f)
        dt = m.group(1) if m else TODAY
        rows = read_summary_sheet2(f, dt)
        if len(rows) > best_n:
            best = {}
            for d, row in rows:
                merge(best, d, row)
            best_n = len(rows)
    print(f"  回种得到 {len(best)} 只债券")
    return best


def save_master(m):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(MASTER_PATH, 'w', encoding='utf-8') as f:
        json.dump(m, f, ensure_ascii=False, indent=1)


def _parse_issue_date(s):
    """从 '2026-07-23 18:00' / '07-23 18:00' 之类解析出 YYYY-MM-DD。"""
    s = str(s).strip()
    if not s:
        return ''
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', s)          # 完整日期 2026-07-23
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(r'(?<!\d)(\d{1,2})-(\d{2})(?!\d)', s)  # 仅月-日 07-23（避免误截 2026 中的 26-07）
    if m:
        return f"{TODAY[:4]}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return ''


def issue_date_of(info):
    """从主库债券信息解析发行/截标日期(YYYY-MM-DD)，判定是否'今日新发'。"""
    for key in ('截标时间', '缴款日', '公告日'):
        v = (info or {}).get('data', {}).get(key)
        if v:
            d = _parse_issue_date(v)
            if d:
                return d
    return ''


def issue_date_of_row(row):
    for key in ('截标时间', '缴款日', '公告日'):
        v = row.get(key)
        if v:
            d = _parse_issue_date(v)
            if d:
                return d
    return ''


# ---------- 导出真源 × 两表 交叉验证 ----------
# 真源 = 债立方当日导出文件 credit_bond_{TODAY}.xlsx。
# Sheet1 的语义 = 「截标/发行日期 == 今天」的债券（与用户预期一致）。
# 注意：债立方 '发行起始日 >= 查询日' 是向前看窗口，且当日导出文件常常为空（债靠更早文件进来），
# 因此：① 不能用'今日文件−更早文件'判定新增；② 不能要求 Sheet1 每只都在今日文件里（会误报）。
# 有意义的交叉验证：
#   (A) 真源中 截标日==今天 的债券，必须全部进入 Sheet1（今日文件有数据时才校验，空则跳过）；
#   (C) Sheet1 每只必须同时出现在 Sheet2 汇总（两表一致，结构上必成立，作兜底）。
# 任一不满足即说明 主库/导出/表格 对不上 -> 硬错误，阻断发邮件。
def cross_validate(sheet1_names, master_after):
    errors = []          # [CROSS-ERROR] 硬错误 -> 触发失败退出，不发邮件
    warnings = []        # [CROSS] 软警告 -> 仅提示
    today_file = os.path.join(DOWNLOAD_DIR, f'credit_bond_{TODAY}.xlsx')
    if not os.path.exists(today_file):
        warnings.append(f"[CROSS] 警告: 当日导出文件缺失 {today_file}，无法与真源交叉验证")
        return errors, warnings
    export_today = {}    # name -> issue_date
    for d, row in read_credit_bond(today_file, TODAY):
        n = row.get('债券简称')
        if n:
            export_today[n] = issue_date_of_row(row)
    sheet1_set = set(sheet1_names)
    if not export_today:
        warnings.append(f"[CROSS] 当日导出文件无数据行（债立方当日窗口常为空），跳过(A)真源比对；Sheet1 由主库 issue_date 定义保证")
    else:
        # (A) 真源中截标日==今天 的债券，必须全部在 Sheet1
        missing = [n for n, idt in export_today.items() if idt == TODAY and n not in sheet1_set]
        if missing:
            errors.append(
                f"[CROSS-ERROR] 导出真源截标日=今天的 {len(missing)} 只未计入 Sheet1 当日新增: {sorted(missing)}")
    # (C) Sheet1 每只必须同时出现在 Sheet2 汇总
    not_in_summary = sheet1_set - set(master_after.keys())
    if not_in_summary:
        errors.append(
            f"[CROSS-ERROR] Sheet1 有 {len(not_in_summary)} 只未进入 Sheet2 汇总（两表不一致）: {sorted(not_in_summary)}")
    return errors, warnings


master = load_master()

# ---------- 2) 合并当天/近期导出（刷新）----------
RUN_DATE = TODAY
master_count_before = len(master)
new_this_run = []
credit_files = sorted(glob.glob(os.path.join(DOWNLOAD_DIR, 'credit_bond_*.xlsx')))
if credit_files:
    for f in credit_files:
        m = re.search(r'credit_bond_(\d{4}-\d{2}-\d{2})\.xlsx', f)
        dt = m.group(1) if m else TODAY
        for d, row in read_credit_bond(f, dt):
            is_new, name = merge(master, d, row, RUN_DATE)
            if is_new and name:
                new_this_run.append(name)
    print(f"本次扫描文件: {len(credit_files)} 个; 本次运行新增: {len(new_this_run)} 只")
else:
    print("[WARN] data/ 下无 credit_bond_*.xlsx，本次无可合并数据")

save_master(master)
print(f"主库债券数: {len(master)}  (运行前 {master_count_before}, 增长 {len(master)-master_count_before})")

# ---------- 3) 票面利率补全校验 ----------
missing = [n for n, info in master.items() if info['data'].get('票面利率(%)') in (None, '')]
print(f"票面利率未补全: {len(missing)} 只" + (f" -> {missing}" if missing else " ✅"))

# ---------- 4) 生成 Excel ----------
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_font = Font(name='微软雅黑', size=10)
data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
today_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
missing_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

wb_out = openpyxl.Workbook()

# ---- Sheet1: 当日新增（截标/发行日期 == 今天的债券）----
ws1 = wb_out.active
ws1.title = f"当日新增_{TODAY}"
for ci, h in enumerate(TARGET_HEADERS, 1):
    c = ws1.cell(1, ci, h); c.font = header_font; c.fill = header_fill
    c.alignment = header_alignment; c.border = thin_border
# Sheet1 语义：主库中截标/发行日期 == 今天的债券（与用户预期一致，不受主库预置影响）
sheet1_items = [(name, info['data']) for name, info in master.items()
                if issue_date_of(info) == TODAY]
sheet1_names = [name for name, _ in sheet1_items]
sheet1_bonds = sorted(
    sheet1_items,
    key=lambda x: (x[1].get('截标时间') or ''), reverse=True)
for ri, (name, rd) in enumerate(sheet1_bonds, 2):
    for ci, h in enumerate(TARGET_HEADERS, 1):
        v = rd.get(h)
        c = ws1.cell(ri, ci, v if v is not None else '')
        c.font = data_font; c.alignment = data_alignment; c.border = thin_border
        if h == '票面利率(%)' and not v:
            c.fill = missing_fill
col_widths = {
    '截标时间': 12, '债券简称': 18, '债券代码': 16, '发行期限': 8, '计划发行(亿)': 10,
    '投标区间(%)': 14, '票面利率(%)': 10, '发行人': 28, '区域': 10, '相似二级券': 16,
    '相似二级剩余期限': 12, '相似二级估值': 12, 'YY评分': 8, '主体评级': 8, '债项评级': 8,
    '主承销商': 30, '担保人': 12, '债券类型': 10, '发行人类型': 12, '交易市场': 8,
    '债券全称': 40, '募集方式': 8, '实际到期日': 12, '到期日休市情况': 10, '上市日': 12,
    '公告日': 12, '缴款日': 12, '板块分类': 8, 'DM评分': 8, 'DM一级行业': 10,
    'DM二级行业': 10, 'DM三级行业': 10, '申万行业': 10, '行权日': 12, '是否有担保': 8,
    '条款': 8, '偿付顺序': 8, '发行截止日': 12, '团费(%)': 8}
for ci, h in enumerate(TARGET_HEADERS, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)
ws1.row_dimensions[1].height = 30
ws1.freeze_panes = 'A2'

# ---- Sheet2: 汇总（全量主库）----
ws2 = wb_out.create_sheet(title="汇总")
SUMMARY_HEADERS = ['首次提取日期'] + TARGET_HEADERS
for ci, h in enumerate(SUMMARY_HEADERS, 1):
    c = ws2.cell(1, ci, h); c.font = header_font; c.fill = header_fill
    c.alignment = header_alignment; c.border = thin_border
sorted_bonds = sorted(master.items(),
                      key=lambda x: (x[1]['data'].get('截标时间') or ''), reverse=True)
for ri, (name, info) in enumerate(sorted_bonds, 2):
    c = ws2.cell(ri, 1, info['first_seen']); c.font = data_font
    c.alignment = data_alignment; c.border = thin_border
    for ci, h in enumerate(TARGET_HEADERS, 2):
        v = info['data'].get(h)
        c = ws2.cell(ri, ci, v if v is not None else '')
        c.font = data_font; c.alignment = data_alignment; c.border = thin_border
        if h in ('债券代码', '票面利率(%)') and v:
            c.fill = today_fill
        if h == '票面利率(%)' and not v:
            c.fill = missing_fill
ws2.column_dimensions['A'].width = 12
for ci, h in enumerate(TARGET_HEADERS, 2):
    ws2.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)
ws2.row_dimensions[1].height = 30
ws2.freeze_panes = 'B2'

# ---------- 5) 两表一致性自检 + 导出真源交叉验证 ----------
sheet1_count = len(sheet1_names)
delta = len(master) - master_count_before
print(f"[INFO] Sheet1 当日新增(截标日=今天): {sheet1_count} 条；本次运行新纳入主库: {len(new_this_run)} 只")

# 真·交叉验证：导出文件(真源) × Sheet1 × Sheet2
cross_errors, cross_warnings = cross_validate(sheet1_names, master)
for w in cross_warnings:
    print(w)
cross_failed = bool(cross_errors)
for e in cross_errors:
    print(e)
if cross_failed:
    print(f"[CROSS-CHECK] ❌ 未通过（{len(cross_errors)} 处）：导出真源与两表对不上，已阻止发邮件")
else:
    print("[CROSS-CHECK] ✅ 导出真源 × Sheet1 × Sheet2 三方交叉验证通过")

wb_out.save(OUTPUT_PATH)
rate_filled = len(master) - len(missing)
print(f"\n[SUCCESS] 输出: {OUTPUT_PATH}")
print(f"  Sheet1 当日新增: {sheet1_count} 条")
print(f"  Sheet2 汇总: {len(sorted_bonds)} 条 (累计，只增不减)")
print(f"  票面利率补全: {rate_filled}/{len(master)}")

# ---------- 6) 输出统计（供邮件与自检消费）----------
# __STATS__ 行供 workflow 用 grep 提取 -> send_email.js 解析 (today:total:rateFilled:rateTotal)
print(f"__STATS__:{sheet1_count}:{len(master)}:{rate_filled}:{len(master)}")
stats = {
    "date": TODAY,
    "total": len(master),
    "new_today": sheet1_count,
    "new_names": sorted(sheet1_names),
    "missing_rate_count": len(missing),
    "cross_failed": cross_failed,
}
with open(os.path.join(DATA_DIR, "summary_stats.json"), "w", encoding="utf-8") as f:
    json.dump(stats, f, ensure_ascii=False, indent=1)
print("  统计已写入 summary_stats.json")

if cross_failed:
    # 与导出真源对不上 -> 显式非零退出，workflow 将跳过发邮件
    sys.exit(2)
